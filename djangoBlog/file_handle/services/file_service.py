'''
Author: qiuzx
Date: 2025-04-09 22:29:00
LastEditors: qiuzx
Description: description
'''
# file_handle/services.py

import hashlib
import os
import re
import zipfile
from django.conf import settings
import pandas as pd
from django.http import FileResponse
from rest_framework.exceptions import APIException
from urllib.parse import quote
from openpyxl import load_workbook
import markdown
from markdown.extensions.toc import TocExtension
from markdown.extensions.codehilite import CodeHiliteExtension
import frontmatter  # 处理YAML front matter
from datetime import datetime
import xlrd  # 处理旧版 Excel 文件
from file_handle.services.libreOffice_service import libreOfficeService
import re
from xml.etree import ElementTree as ET
import json
from xmindparser import xmind_to_dict
from file_handle.services.project_service import ProjectService


class FileService:
    @staticmethod
    def get_cache_path(file_path):
        cache_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'blog_temp_data', 'pdf_cache')
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir)
        # 使用文件路径和最后修改时间生成唯一的缓存文件名
        file_stat = os.stat(file_path)
        cache_key = f"{file_path}_{file_stat.st_mtime}"
        cache_name = hashlib.md5(cache_key.encode()).hexdigest() + '.pdf'
        return os.path.join(cache_dir, cache_name)
    
    @staticmethod
    def fix_image_paths(md_content, md_file_path, base_url):
        """
        修复 Markdown 文件中的图片路径
        :param md_content: Markdown 内容
        :param md_file_path: Markdown 文件路径
        :param base_url: 静态文件基础URL
        """
        # 获取系统配置的文档库路径
        doc_root = ProjectService.get_system_path()
        if not doc_root:
            raise APIException('系统配置路径不存在，请先配置系统路径')
        # 从配置路径中提取最后一段作为 URL 前缀
        url_prefix = os.path.basename(doc_root)
        # 计算相对路径
        relative_dir = os.path.relpath(os.path.dirname(md_file_path), doc_root)
        def replace_src(match):
            src_value = match.group(1)
            if src_value.startswith('http'):
                return match.group(0)
            print(base_url,'1212121')
            full_url = f"{base_url}/{url_prefix}/{relative_dir.rstrip('/')}/{src_value.lstrip('/')}"
            return f'src="{full_url}"'
        return re.sub(r'src="([^"]+)"', replace_src, md_content)

    @staticmethod
    def read_file_content(file_path, file_type):
        try:
            if file_type in ['doc', 'docx', 'ppt', 'pptx']:
                # 检查缓存
                return libreOfficeService.convert_to_pdf(file_path) 
            elif file_type == 'xlsx':
                # 处理新版 Excel 文件 (.xlsx)
                sheets_data = []
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_data = {
                        'name': sheet_name,
                        'headers': [],
                        'rows': []
                    }
                    
                    # 获取所有列的最大宽度
                    max_lengths = {}
                    for col in sheet.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                max_length = max(max_length, len(str(cell.value or '')))
                            except:
                                pass
                        max_lengths[col_letter] = max_length
                    
                    # 获取表头
                    headers = []
                    first_row = next(sheet.rows)
                    for cell in first_row:
                        headers.append({
                            'value': str(cell.value) if cell.value is not None else '',
                            'width': max_lengths[cell.column_letter]
                        })
                    sheet_data['headers'] = headers
                    
                    # 获取数据行
                    for row in list(sheet.rows)[1:]:
                        row_data = []
                        for cell in row:
                            value = cell.value
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row_data.append(str(value))
                        sheet_data['rows'].append(row_data)
                    
                    sheets_data.append(sheet_data)
                return {
                    'content': sheets_data,
                    'type': 'xlsx',
                    'meta': {
                        'sheets_count': len(sheets_data),
                        'filename': os.path.basename(file_path)
                    }
                }
            elif file_type == 'xls':
                # 处理旧版 Excel 文件 (.xls)
                sheets_data = []
                wb = xlrd.open_workbook(file_path)
                for sheet_name in wb.sheet_names():
                    sheet = wb.sheet_by_name(sheet_name)
                    sheet_data = {
                        'name': sheet_name,
                        'headers': [],
                        'rows': []
                    }
                    
                    # 获取所有列的最大宽度
                    max_lengths = [0] * sheet.ncols
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            try:
                                cell_value = str(sheet.cell_value(row_idx, col_idx))
                                max_lengths[col_idx] = max(max_lengths[col_idx], len(cell_value))
                            except:
                                pass
                    
                    # 获取表头
                    headers = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(0, col_idx)
                        headers.append({
                            'value': str(cell_value) if cell_value else '',
                            'width': max_lengths[col_idx]
                        })
                    sheet_data['headers'] = headers
                    
                    # 获取数据行
                    for row_idx in range(1, sheet.nrows):
                        row_data = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            value = cell.value
                            
                            # 处理日期类型
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    # Excel 中的日期是从 1900-01-01 开始的天数
                                    value = datetime(*xlrd.xldate_as_tuple(value, wb.datemode))
                                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                                except:
                                    value = str(value)
                            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                                # 处理数字，避免显示科学计数法
                                if value.is_integer():
                                    value = int(value)
                                value = str(value)
                            else:
                                value = str(value) if value else ''
                                
                            row_data.append(value)
                        sheet_data['rows'].append(row_data)
                    
                    sheets_data.append(sheet_data)
            
                return {
                    'content': sheets_data,
                    'type': 'xls',
                    'meta': {
                        'sheets_count': len(sheets_data),
                        'filename': os.path.basename(file_path)
                    }
                }
            elif file_type == 'pdf':
                # 返回 PDF 文件流
                response = FileResponse(open(file_path, 'rb'), content_type='application/pdf')
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"; filename*=UTF-8\'\'{quote(os.path.basename(file_path))}'
                return response
            elif file_type == 'md':
                with open(file_path, 'r', encoding='utf-8') as f:
                    try:
                        post = frontmatter.load(f)
                        fixed_content = FileService.fix_image_paths(
                            post.content,
                            file_path,
                            f'{settings.FILE_HANDLE_API_BASE}'   # 使用配置的URL前缀
                        )
                        return {
                            'content': fixed_content,  # 原始Markdown文本
                            'meta': post.metadata,    # Front Matter元数据
                            'type': 'md'              # 标记为Markdown类型
                        }
                    except Exception as e:
                        raise APIException(f"Error parsing Markdown file: {str(e)}")
            elif file_type == 'txt' or file_type == 'etf':
                encodings = ['utf-8', 'gbk', 'iso-8859-1']
                for enc in encodings:
                    try:
                        with open(file_path, 'r', encoding=enc) as f:
                            return {
                                'content': f.read(),
                                'type': 'txt',
                                'encoding': enc  # 可选：返回使用的编码
                            }
                    except UnicodeDecodeError:
                        continue
                raise ValueError("Failed to decode file using common encodings.")
            elif file_type == 'png' or file_type == 'jpg':
                return FileResponse(open(file_path, 'rb'), content_type='image/png')
            elif file_type == 'xmind':
                data = xmind_to_dict(file_path)
                return data
            else:
                raise APIException(f"Error reading xmind file: {str(e)}")   
        except Exception as e:
            raise APIException(f"Error reading file: {str(e)}")
