"""
文件服务层 - 对应 file_handle/services/file_service.py
"""
import os
import re
from datetime import datetime
from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
import frontmatter
import xlrd
from xmindparser import xmind_to_dict
from app.config import settings
from app.file_handle.services.libreoffice_service import LibreOfficeService


class FileService:

    @staticmethod
    def fix_image_paths(md_content: str, md_file_path: str, base_url: str, doc_root: str) -> str:
        """
        修复 Markdown 文件中的图片路径
        """
        if not doc_root:
            raise HTTPException(status_code=400, detail="系统配置路径不存在，请先配置系统路径")

        url_prefix = os.path.basename(doc_root)
        relative_dir = os.path.relpath(os.path.dirname(md_file_path), doc_root)

        def replace_src(match):
            src_value = match.group(1)
            if src_value.startswith('http'):
                return match.group(0)
            full_url = f"{base_url}/{url_prefix}/{relative_dir.rstrip('/')}/{src_value.lstrip('/')}"
            return f'src="{full_url}"'

        return re.sub(r'src="([^"]+)"', replace_src, md_content)

    @staticmethod
    def read_file_content(file_path: str, file_type: str, doc_root: str = ""):
        """读取文件内容，根据文件类型返回不同格式"""
        try:
            if file_type in ['doc', 'docx', 'ppt', 'pptx']:
                return LibreOfficeService.convert_to_pdf(file_path)

            elif file_type == 'xlsx':
                sheets_data = []
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    sheet_data = {
                        'name': sheet_name,
                        'headers': [],
                        'rows': []
                    }

                    max_lengths = {}
                    for col in sheet.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                max_length = max(max_length, len(str(cell.value or '')))
                            except:
                                pass
                        max_lengths[col_letter] = max_length

                    headers = []
                    first_row = next(sheet.rows)
                    for cell in first_row:
                        headers.append({
                            'value': str(cell.value) if cell.value is not None else '',
                            'width': max_lengths[cell.column_letter]
                        })
                    sheet_data['headers'] = headers

                    for row in list(sheet.rows)[1:]:
                        row_data = []
                        for cell in row:
                            value = cell.value
                            if isinstance(value, datetime):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = ''
                            row_data.append(str(value))
                        sheet_data['rows'].append(row_data)

                    sheets_data.append(sheet_data)

                return {
                    'content': sheets_data,
                    'type': 'xlsx',
                    'meta': {
                        'sheets_count': len(sheets_data),
                        'filename': os.path.basename(file_path)
                    }
                }

            elif file_type == 'xls':
                sheets_data = []
                wb = xlrd.open_workbook(file_path)
                for sheet_name in wb.sheet_names():
                    sheet = wb.sheet_by_name(sheet_name)
                    sheet_data = {
                        'name': sheet_name,
                        'headers': [],
                        'rows': []
                    }

                    max_lengths = [0] * sheet.ncols
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            try:
                                cell_value = str(sheet.cell_value(row_idx, col_idx))
                                max_lengths[col_idx] = max(max_lengths[col_idx], len(cell_value))
                            except:
                                pass

                    headers = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(0, col_idx)
                        headers.append({
                            'value': str(cell_value) if cell_value else '',
                            'width': max_lengths[col_idx]
                        })
                    sheet_data['headers'] = headers

                    for row_idx in range(1, sheet.nrows):
                        row_data = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            value = cell.value

                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    value = datetime(*xlrd.xldate_as_tuple(value, wb.datemode))
                                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                                except:
                                    value = str(value)
                            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                                if value == int(value):
                                    value = int(value)
                                value = str(value)
                            else:
                                value = str(value) if value else ''

                            row_data.append(value)
                        sheet_data['rows'].append(row_data)

                    sheets_data.append(sheet_data)

                return {
                    'content': sheets_data,
                    'type': 'xls',
                    'meta': {
                        'sheets_count': len(sheets_data),
                        'filename': os.path.basename(file_path)
                    }
                }

            elif file_type == 'pdf':
                return FileResponse(
                    path=file_path,
                    media_type='application/pdf',
                    filename=os.path.basename(file_path),
                    headers={
                        'Content-Disposition': f'inline; filename="{os.path.basename(file_path)}"'
                    }
                )

            elif file_type == 'md':
                with open(file_path, 'r', encoding='utf-8') as f:
                    try:
                        post = frontmatter.load(f)
                        fixed_content = FileService.fix_image_paths(
                            post.content,
                            file_path,
                            settings.FILE_HANDLE_API_BASE,
                            doc_root
                        )
                        return {
                            'content': fixed_content,
                            'meta': post.metadata,
                            'type': 'md'
                        }
                    except Exception as e:
                        raise HTTPException(
                            status_code=500,
                            detail=f"Error parsing Markdown file: {str(e)}"
                        )

            elif file_type in ['txt', 'etf']:
                encodings = ['utf-8', 'gbk', 'iso-8859-1']
                for enc in encodings:
                    try:
                        with open(file_path, 'r', encoding=enc) as f:
                            return {
                                'content': f.read(),
                                'type': 'txt',
                                'encoding': enc
                            }
                    except UnicodeDecodeError:
                        continue
                raise HTTPException(status_code=500, detail="Failed to decode file using common encodings.")

            elif file_type in ['png', 'jpg']:
                media_type = 'image/png' if file_type == 'png' else 'image/jpeg'
                return FileResponse(path=file_path, media_type=media_type)

            elif file_type == 'xmind':
                data = xmind_to_dict(file_path)
                return data

            else:
                raise HTTPException(status_code=400, detail=f"不支持的文件类型: {file_type}")

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error reading file: {str(e)}")
